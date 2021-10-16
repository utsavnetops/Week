from extract import closedcases
from chg_extract import chgext
from actcase import actcase
from json import loads
from openpyxl import load_workbook
from datetime import datetime
# This is a sample Python script.
#C:\\Users\\utsakuma\\Desktop\\scrypt\\
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def date():
    start = input("Start date in format(DDMMYYYY): ")
    datetime_obj = datetime.strptime(start, "%d%m%Y")
    datetime_obj = datetime_obj.toordinal()
    derange = [datetime.fromordinal(datetime_obj + dates).date().strftime("%Y-%m-%d") for dates in range(0, 7, 1)]
    print (derange)
    #derange = [datetime.fromordinal(datetime_obj+dates).date() for dates in range(1,8,1)]
    #end = input("End date in two digit format(XX): ")
    #month = input("Month in two digit format(XX): ")
    #year = input("Year in four digit format(XXXX): ")
    #my = year + "-" + month + "-"
    #derange = ["2021-05-30","2021-05-31","2021-06-01","2021-06-02","2021-06-03","2021-06-04","2021-06-05"]
    #derange = [ my + str(x) for x in range(int(start), int(end) + 1)]
    return derange


def fname():
    fnamecase = input("Name of case file: ")
    fnamechg1 = input("Name of 1st change file: ")
    fnamechg2 = input("Name of 2nd change file: ")
    fnamedest = input("Destination file name: ")
    fullpath = "C:\\Users\\utsakuma\\Desktop\\scrypt\\"
    fnames = [ fullpath+name for name in [fnamecase,fnamechg1,fnamechg2,fnamedest]]
    return fnames


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print("This program is to get one week report")
    filenames = fname()
    rdate = date()
    with open("employee.txt", 'r', encoding='utf-8') as f:
        empdump = f.readline()

    employees = loads(empdump)

    for ids in employees:
        print(ids,end=', ')
    print()
    load_wbs = [load_workbook(filename) for filename in filenames]
    load_sheets = [sheet.worksheets[0] for sheet in load_wbs]
    load_sheets.append(load_wbs[3].create_sheet('Active'))
    load_sheets.append(load_wbs[3].create_sheet('Change'))

    closedcases(load_sheets[0],load_sheets[3],rdate,employees)
    chgext(load_sheets[1],load_sheets[2],load_sheets[5],rdate,employees)
    actcase(load_sheets[0],load_sheets[4],employees)

    load_wbs[3].save(filenames[3])

# See PyCharm help at https://www.jetbrains.com/help/pycharm/